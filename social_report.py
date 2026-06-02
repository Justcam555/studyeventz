#!/usr/bin/env python3
"""
Social Media Report Generator for AU University Agents

Generates analysis and content ideas for social media based on agent database.

Usage:
    python social_report.py                          # Full report (HTML + Excel)
    python social_report.py --country "China"        # Country-specific report
    python social_report.py --university "Monash"    # University-specific report
    python social_report.py --format excel           # Excel only
    python social_report.py --format html            # HTML only
    python social_report.py --output ./reports/      # Output directory
"""

import argparse
import sqlite3
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

DB_PATH = Path(__file__).parent / "data" / "agents.db"
REPORTS_DIR = Path(__file__).parent / "reports"


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ─── Data gathering ───────────────────────────────────────────────────────────

def gather_data(conn, country=None, university=None):
    filters = []
    params = []
    if country:
        filters.append("LOWER(a.country) LIKE LOWER(?)")
        params.append(f"%{country}%")
    if university:
        filters.append("LOWER(u.name) LIKE LOWER(?)")
        params.append(f"%{university}%")
    where = ("WHERE " + " AND ".join(filters)) if filters else ""

    # Summary totals
    summary = conn.execute(f"""
        SELECT
            COUNT(DISTINCT a.id)        AS total_agents,
            COUNT(DISTINCT a.country)   AS countries,
            COUNT(DISTINCT u.id)        AS universities,
            COUNT(a.email)              AS with_email,
            COUNT(a.website)            AS with_website,
            COUNT(a.phone)              AS with_phone
        FROM agents a JOIN universities u ON u.id = a.university_id
        {where}
    """, params).fetchone()

    # By country
    by_country = pd.read_sql_query(f"""
        SELECT
            COALESCE(a.country,'Unknown') AS country,
            COUNT(*)                      AS agents,
            COUNT(DISTINCT u.id)          AS universities
        FROM agents a JOIN universities u ON u.id = a.university_id
        {where}
        GROUP BY 1 ORDER BY 2 DESC LIMIT 30
    """, conn, params=params)

    # By university
    by_uni = pd.read_sql_query(f"""
        SELECT
            u.name      AS university,
            COUNT(a.id) AS agents,
            COUNT(DISTINCT a.country) AS countries
        FROM universities u
        LEFT JOIN agents a ON a.university_id = u.id
        {'JOIN agents a2 ON a2.university_id = u.id ' + where.replace('a.','a2.') if where else ''}
        GROUP BY u.id ORDER BY 2 DESC
    """, conn, params=params if where else [])

    # Top agents (appearing at multiple unis)
    multi_uni = pd.read_sql_query(f"""
        SELECT
            a.company_name,
            COUNT(DISTINCT u.id)        AS university_count,
            GROUP_CONCAT(DISTINCT u.name) AS universities,
            COUNT(DISTINCT a.country)   AS countries,
            a.email, a.website
        FROM agents a JOIN universities u ON u.id = a.university_id
        {where}
        GROUP BY LOWER(TRIM(a.company_name))
        HAVING university_count > 1
        ORDER BY university_count DESC, countries DESC
        LIMIT 50
    """, conn, params=params)

    # Recent scrapes
    scrape_log = pd.read_sql_query("""
        SELECT u.name, sl.scraped_at, sl.status, sl.agents_found, sl.method
        FROM scrape_log sl JOIN universities u ON u.id = sl.university_id
        ORDER BY sl.scraped_at DESC LIMIT 50
    """, conn)

    return {
        "summary": dict(summary),
        "by_country": by_country,
        "by_university": by_uni,
        "multi_university_agents": multi_uni,
        "scrape_log": scrape_log,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "filter_country": country,
        "filter_university": university,
    }


# ─── HTML Report ──────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>AU University Agents — Social Media Report</title>
<style>
  :root {{
    --primary: #003366;
    --accent:  #F4A300;
    --light:   #F0F4F8;
    --text:    #1A1A2E;
    --border:  #CBD5E1;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #F8FAFC; color: var(--text); }}
  .header {{
    background: linear-gradient(135deg, var(--primary) 0%, #005A9E 100%);
    color: #fff; padding: 2.5rem 3rem;
  }}
  .header h1 {{ font-size: 2rem; margin-bottom: .5rem; }}
  .header p {{ opacity: .8; font-size: .95rem; }}
  .container {{ max-width: 1300px; margin: 0 auto; padding: 2rem 2.5rem; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 1rem; margin-bottom: 2.5rem; }}
  .kpi-card {{
    background: #fff; border-radius: 12px; padding: 1.5rem 1rem;
    text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,.06);
    border-top: 4px solid var(--accent);
  }}
  .kpi-card .value {{ font-size: 2.2rem; font-weight: 700; color: var(--primary); }}
  .kpi-card .label {{ font-size: .8rem; color: #64748B; margin-top: .3rem; text-transform: uppercase; letter-spacing: .05em; }}
  .section {{ background: #fff; border-radius: 12px; padding: 1.5rem; margin-bottom: 2rem; box-shadow: 0 2px 8px rgba(0,0,0,.06); }}
  .section h2 {{ font-size: 1.1rem; color: var(--primary); margin-bottom: 1rem; padding-bottom: .5rem; border-bottom: 2px solid var(--light); }}
  table {{ width: 100%; border-collapse: collapse; font-size: .88rem; }}
  th {{ background: var(--primary); color: #fff; padding: .6rem .8rem; text-align: left; font-weight: 600; }}
  td {{ padding: .5rem .8rem; border-bottom: 1px solid var(--border); }}
  tr:nth-child(even) td {{ background: var(--light); }}
  tr:hover td {{ background: #E8F0FE; }}
  .bar-wrap {{ display: flex; align-items: center; gap: .5rem; }}
  .bar {{ height: 14px; background: var(--accent); border-radius: 3px; min-width: 2px; transition: width .3s; }}
  .bar-val {{ font-size: .8rem; color: #64748B; white-space: nowrap; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: .75rem; font-weight: 600; }}
  .badge-green {{ background: #D1FAE5; color: #065F46; }}
  .badge-blue  {{ background: #DBEAFE; color: #1E40AF; }}
  .badge-amber {{ background: #FEF3C7; color: #92400E; }}
  .social-card {{
    border: 1px solid var(--border); border-radius: 10px; padding: 1rem 1.2rem;
    margin-bottom: .8rem; position: relative;
  }}
  .social-card .platform {{ font-size: .75rem; font-weight: 700; color: var(--accent); text-transform: uppercase; margin-bottom: .4rem; }}
  .social-card p {{ line-height: 1.6; font-size: .9rem; }}
  .social-card .hashtags {{ margin-top: .5rem; color: #2563EB; font-size: .85rem; }}
  .copy-btn {{
    position: absolute; top: .8rem; right: .8rem;
    background: var(--light); border: 1px solid var(--border);
    border-radius: 6px; padding: 3px 10px; font-size: .75rem;
    cursor: pointer; color: var(--text);
  }}
  .copy-btn:hover {{ background: var(--border); }}
  footer {{ text-align: center; padding: 2rem; color: #94A3B8; font-size: .8rem; }}
</style>
</head>
<body>

<div class="header">
  <h1>🎓 Australian University Agent Network</h1>
  <p>Social Media Intelligence Report &nbsp;·&nbsp; Generated {generated_at}
  {filter_label}
  </p>
</div>

<div class="container">

  <!-- KPIs -->
  <div class="kpi-grid">
    <div class="kpi-card"><div class="value">{total_agents}</div><div class="label">Total Agents</div></div>
    <div class="kpi-card"><div class="value">{countries}</div><div class="label">Countries</div></div>
    <div class="kpi-card"><div class="value">{universities}</div><div class="label">Universities</div></div>
    <div class="kpi-card"><div class="value">{with_email}</div><div class="label">With Email</div></div>
    <div class="kpi-card"><div class="value">{with_website}</div><div class="label">With Website</div></div>
    <div class="kpi-card"><div class="value">{email_pct}%</div><div class="label">Email Coverage</div></div>
  </div>

  <!-- By Country -->
  <div class="section">
    <h2>📍 Agent Distribution by Country (Top 30)</h2>
    <table>
      <thead><tr><th>Country</th><th>Agents</th><th>Universities</th><th>Volume</th></tr></thead>
      <tbody>
        {country_rows}
      </tbody>
    </table>
  </div>

  <!-- By University -->
  <div class="section">
    <h2>🏛️ Agent Count by University</h2>
    <table>
      <thead><tr><th>University</th><th>Agents</th><th>Countries</th><th>Coverage</th></tr></thead>
      <tbody>
        {uni_rows}
      </tbody>
    </table>
  </div>

  <!-- Multi-University Agents -->
  <div class="section">
    <h2>🔗 Agents Representing Multiple Universities</h2>
    <p style="font-size:.85rem;color:#64748B;margin-bottom:1rem;">
      These agencies are authorised by more than one Australian university — high-value partnership targets.
    </p>
    <table>
      <thead><tr><th>Agency</th><th>Unis</th><th>Countries</th><th>Email</th><th>Website</th></tr></thead>
      <tbody>
        {multi_rows}
      </tbody>
    </table>
  </div>

  <!-- Social Media Content Ideas -->
  <div class="section">
    <h2>📱 Social Media Content Ideas</h2>
    {social_cards}
  </div>

</div>

<footer>Australian University Agent Database &nbsp;·&nbsp; {generated_at}</footer>

<script>
document.querySelectorAll('.copy-btn').forEach(btn => {{
  btn.addEventListener('click', () => {{
    const card = btn.closest('.social-card');
    const text = card.querySelector('p').innerText + '\\n' + (card.querySelector('.hashtags')?.innerText || '');
    navigator.clipboard.writeText(text).then(() => {{
      btn.textContent = 'Copied!';
      setTimeout(() => btn.textContent = 'Copy', 1500);
    }});
  }});
}});
</script>
</body>
</html>
"""


def build_html(data):
    s = data["summary"]
    email_pct = round(100 * s["with_email"] / s["total_agents"]) if s["total_agents"] else 0

    filter_label = ""
    if data["filter_country"]:
        filter_label = f" &nbsp;·&nbsp; Country: <strong>{data['filter_country']}</strong>"
    if data["filter_university"]:
        filter_label += f" &nbsp;·&nbsp; University: <strong>{data['filter_university']}</strong>"

    # Country rows
    max_agents = data["by_country"]["agents"].max() if not data["by_country"].empty else 1
    country_rows = ""
    for _, r in data["by_country"].iterrows():
        bar_w = int(200 * r["agents"] / max_agents)
        country_rows += f"""
        <tr>
          <td>{r['country']}</td>
          <td><strong>{r['agents']}</strong></td>
          <td>{r['universities']}</td>
          <td><div class="bar-wrap"><div class="bar" style="width:{bar_w}px"></div><span class="bar-val">{r['agents']}</span></div></td>
        </tr>"""

    # University rows
    uni_rows = ""
    max_u = data["by_university"]["agents"].max() if not data["by_university"].empty else 1
    for _, r in data["by_university"].iterrows():
        bar_w = int(150 * r["agents"] / max_u) if max_u > 0 else 0
        badge = '<span class="badge badge-green">✓ Data</span>' if r["agents"] > 0 else '<span class="badge badge-amber">No data</span>'
        uni_rows += f"""
        <tr>
          <td>{r['university']}</td>
          <td><strong>{int(r['agents'])}</strong></td>
          <td>{int(r['countries']) if r['countries'] else 0}</td>
          <td>{badge}</td>
        </tr>"""

    # Multi-uni rows
    multi_rows = ""
    for _, r in data["multi_university_agents"].iterrows():
        email_td = f'<a href="mailto:{r["email"]}">{r["email"]}</a>' if r["email"] else "—"
        web_td = f'<a href="{r["website"]}" target="_blank">🔗</a>' if r["website"] else "—"
        multi_rows += f"""
        <tr>
          <td><strong>{r['company_name']}</strong></td>
          <td><span class="badge badge-blue">{int(r['university_count'])}</span></td>
          <td>{int(r['countries'])}</td>
          <td>{email_td}</td>
          <td>{web_td}</td>
        </tr>"""
    if not multi_rows:
        multi_rows = "<tr><td colspan='5' style='text-align:center;color:#94A3B8;padding:1.5rem'>No multi-university agents found yet (more scraping needed)</td></tr>"

    # Social media posts
    top_countries = data["by_country"].head(5)["country"].tolist() if not data["by_country"].empty else []
    country_list = ", ".join(top_countries[:5]) if top_countries else "multiple countries"
    posts = [
        {
            "platform": "LinkedIn",
            "text": f"Australia's universities have a global network of {s['total_agents']:,} authorised education agents across {s['countries']} countries. These trusted partners are the bridge connecting international students with world-class Australian education. From {country_list} and beyond — the reach is remarkable. 🌏",
            "hashtags": "#StudyInAustralia #InternationalEducation #EducationAgents #Australia"
        },
        {
            "platform": "Twitter / X",
            "text": f"Did you know? Australian universities work with {s['total_agents']:,} authorised education agents in {s['countries']} countries 📊 That's a massive global network helping students find the right study pathway 🇦🇺",
            "hashtags": "#StudyInAustralia #HigherEducation #InternationalStudents"
        },
        {
            "platform": "Instagram Caption",
            "text": f"Behind every international student's journey to Australia is an authorised education agent ✈️🎓 Our analysis of Australia's {s['universities']} universities reveals a network of {s['total_agents']:,} agents spanning {s['countries']} countries worldwide. Whether you're in {top_countries[0] if top_countries else 'Asia'}, Europe, or the Americas — there's a trusted expert near you to guide your study abroad adventure.",
            "hashtags": "#StudyAbroad #Australia #StudyInAustralia #InternationalStudent #UniversityLife #AustralianEducation #GlobalEducation"
        },
        {
            "platform": "Facebook",
            "text": f"Thinking about studying in Australia? 🦘\n\nAustralia's top universities maintain a carefully vetted network of {s['total_agents']:,} authorised education agents across {s['countries']} countries — making quality guidance accessible wherever you are.\n\nThese agents are approved directly by universities, giving you peace of mind that you're getting genuine, accurate advice about your study options.\n\nReach out to find an authorised agent near you! 👇",
            "hashtags": "#StudyInAustralia #AustralianUniversities #InternationalEducation"
        },
        {
            "platform": "LinkedIn (Data Insight)",
            "text": f"📊 Agent Network Data Insight\n\nOur analysis of publicly listed education agents across Australian universities reveals:\n• {s['total_agents']:,} total authorised agents\n• {s['countries']} countries represented\n• {email_pct}% of agents have publicly listed contact details\n\nFor institutions looking to expand international recruitment, the agent channel remains the dominant pathway for international student acquisition in Australian higher education.",
            "hashtags": "#HigherEducation #InternationalRecruitment #EdTech #StudentRecruitment #Australia"
        },
    ]

    social_cards = ""
    for post in posts:
        social_cards += f"""
        <div class="social-card">
          <button class="copy-btn">Copy</button>
          <div class="platform">{post['platform']}</div>
          <p>{post['text'].replace(chr(10), '<br>')}</p>
          <div class="hashtags">{post['hashtags']}</div>
        </div>"""

    return HTML_TEMPLATE.format(
        generated_at=data["generated_at"],
        filter_label=filter_label,
        total_agents=f"{s['total_agents']:,}",
        countries=s["countries"],
        universities=s["universities"],
        with_email=f"{s['with_email']:,}",
        with_website=f"{s['with_website']:,}",
        email_pct=email_pct,
        country_rows=country_rows,
        uni_rows=uni_rows,
        multi_rows=multi_rows,
        social_cards=social_cards,
    )


# ─── Excel Report ─────────────────────────────────────────────────────────────

BLUE  = "1F4E79"
AMBER = "F4A300"
WHITE = "FFFFFF"
LIGHT = "DCE6F1"


def _hdr(ws, row, col, value, bg=BLUE, fg=WHITE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=fg, name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    return c


def _cell(ws, row, col, value, bold=False, link=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Arial", size=9,
                  color="0563C1" if link else "1A1A2E",
                  underline="single" if link else "none")
    if link:
        c.hyperlink = link
    if row % 2 == 0:
        c.fill = PatternFill("solid", start_color=LIGHT)
    return c


def build_excel(data, path):
    wb = Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Australian University Agent Network"
    ws["A1"].font = Font(bold=True, size=16, color=BLUE, name="Arial")
    ws["A2"] = f"Report generated: {data['generated_at']}"
    ws["A2"].font = Font(size=10, color="64748B", name="Arial")
    ws.row_dimensions[1].height = 28

    s = data["summary"]
    email_pct = round(100 * s["with_email"] / s["total_agents"]) if s["total_agents"] else 0

    kpis = [
        ("Total Agents", s["total_agents"]),
        ("Countries", s["countries"]),
        ("Universities", s["universities"]),
        ("With Email", s["with_email"]),
        ("With Website", s["with_website"]),
        ("Email Coverage %", f"{email_pct}%"),
    ]
    for ci, (label, val) in enumerate(kpis, 1):
        _hdr(ws, 4, ci, label, AMBER, BLUE)
        c = ws.cell(row=5, column=ci, value=val)
        c.font = Font(bold=True, size=14, name="Arial", color=BLUE)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # ── Sheet 2: By Country ────────────────────────────────────────────
    ws2 = wb.create_sheet("By Country")
    ws2.sheet_view.showGridLines = False
    for ci, h in enumerate(["Country", "Agents", "Universities"], 1):
        _hdr(ws2, 1, ci, h)
    for ri, r in enumerate(data["by_country"].itertuples(index=False), 2):
        _cell(ws2, ri, 1, r.country)
        _cell(ws2, ri, 2, r.agents, bold=True)
        _cell(ws2, ri, 3, r.universities)

    # Add bar chart
    if len(data["by_country"]) > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Agents by Country (Top 30)"
        chart.style = 10
        chart.y_axis.title = "Country"
        chart.x_axis.title = "Agents"
        data_ref = Reference(ws2, min_col=2, max_col=2, min_row=1, max_row=min(31, len(data["by_country"]) + 1))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=min(31, len(data["by_country"]) + 1))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.height = 20
        chart.width = 20
        ws2.add_chart(chart, "E2")

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 25 if col == "A" else 12

    # ── Sheet 3: By University ─────────────────────────────────────────
    ws3 = wb.create_sheet("By University")
    ws3.sheet_view.showGridLines = False
    for ci, h in enumerate(["University", "Agents", "Countries Covered"], 1):
        _hdr(ws3, 1, ci, h)
    for ri, r in enumerate(data["by_university"].itertuples(index=False), 2):
        _cell(ws3, ri, 1, r.university)
        _cell(ws3, ri, 2, int(r.agents) if r.agents else 0, bold=True)
        _cell(ws3, ri, 3, int(r.countries) if r.countries else 0)
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 18

    # ── Sheet 4: Multi-University Agents ──────────────────────────────
    ws4 = wb.create_sheet("Multi-Uni Agents")
    ws4.sheet_view.showGridLines = False
    for ci, h in enumerate(["Agency", "Uni Count", "Countries", "Email", "Website", "Universities"], 1):
        _hdr(ws4, 1, ci, h)
    for ri, r in enumerate(data["multi_university_agents"].itertuples(index=False), 2):
        _cell(ws4, ri, 1, r.company_name, bold=True)
        _cell(ws4, ri, 2, int(r.university_count))
        _cell(ws4, ri, 3, int(r.countries))
        _cell(ws4, ri, 4, r.email, link=f"mailto:{r.email}" if r.email else None)
        _cell(ws4, ri, 5, r.website, link=r.website if r.website else None)
        _cell(ws4, ri, 6, r.universities)
    for col, w in zip(["A","B","C","D","E","F"], [40,10,10,35,35,60]):
        ws4.column_dimensions[col].width = w

    # ── Sheet 5: Social Media Posts ────────────────────────────────────
    ws5 = wb.create_sheet("Social Media Posts")
    ws5.sheet_view.showGridLines = False

    posts = [
        ("LinkedIn", f"Australia's universities have a global network of {s['total_agents']:,} authorised education agents across {s['countries']} countries. These trusted partners are the bridge connecting international students with world-class Australian education. 🌏\n\n#StudyInAustralia #InternationalEducation #EducationAgents"),
        ("Twitter/X", f"Did you know? Australian universities work with {s['total_agents']:,} authorised education agents in {s['countries']} countries 📊 That's a massive global network helping students find the right study pathway 🇦🇺 #StudyInAustralia #HigherEducation"),
        ("Instagram", f"Behind every international student's journey to Australia is an authorised education agent ✈️🎓\n\n{s['total_agents']:,} agents. {s['countries']} countries. {s['universities']} universities.\n\nYour study abroad journey starts with the right advice. 🌟\n\n#StudyAbroad #Australia #StudyInAustralia #InternationalStudent #UniversityLife"),
        ("Facebook",  f"Thinking about studying in Australia? 🦘\n\nAustralia's top universities maintain a carefully vetted network of {s['total_agents']:,} authorised education agents across {s['countries']} countries.\n\nReach out to find an authorised agent near you! 👇\n\n#StudyInAustralia #AustralianUniversities #InternationalEducation"),
        ("LinkedIn (Data)", f"📊 Agent Network Data Insight\n\nOur analysis of publicly listed education agents:\n• {s['total_agents']:,} total authorised agents\n• {s['countries']} countries represented\n• {email_pct}% have public contact details\n\n#HigherEducation #InternationalRecruitment #Australia"),
    ]

    _hdr(ws5, 1, 1, "Platform")
    _hdr(ws5, 1, 2, "Post Content (ready to use)")
    for ri, (platform, content) in enumerate(posts, 2):
        c_plat = ws5.cell(row=ri, column=1, value=platform)
        c_plat.font = Font(bold=True, name="Arial", size=9, color=BLUE)
        c_plat.fill = PatternFill("solid", start_color=LIGHT)
        c_cont = ws5.cell(row=ri, column=2, value=content)
        c_cont.font = Font(name="Arial", size=9)
        c_cont.alignment = Alignment(wrap_text=True)
        ws5.row_dimensions[ri].height = 80

    ws5.column_dimensions["A"].width = 18
    ws5.column_dimensions["B"].width = 100

    wb.save(path)
    print(f"✅  Excel report saved → {path}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Social Media Report Generator")
    parser.add_argument("--country",    help="Filter by country")
    parser.add_argument("--university", help="Filter by university")
    parser.add_argument("--format",     choices=["html", "excel", "both"], default="both")
    parser.add_argument("--output",     default=str(REPORTS_DIR), help="Output directory")
    args = parser.parse_args()

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    conn = get_conn()
    data = gather_data(conn, country=args.country, university=args.university)
    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    suffix = f"_{args.country}" if args.country else ""
    suffix += f"_{args.university.replace(' ','_')}" if args.university else ""

    print(f"\n📊 Report Summary:")
    print(f"   Total agents:  {data['summary']['total_agents']:,}")
    print(f"   Countries:     {data['summary']['countries']}")
    print(f"   Universities:  {data['summary']['universities']}")
    print(f"   With email:    {data['summary']['with_email']:,}")

    if args.format in ("html", "both"):
        html_path = out_dir / f"agent_report{suffix}_{ts}.html"
        html_path.write_text(build_html(data), encoding="utf-8")
        print(f"✅  HTML report → {html_path}")

    if args.format in ("excel", "both"):
        excel_path = out_dir / f"agent_report{suffix}_{ts}.xlsx"
        build_excel(data, excel_path)


if __name__ == "__main__":
    main()
