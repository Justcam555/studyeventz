#!/usr/bin/env python3
"""
process_market_data.py — Parse AEI + Home Affairs files into per-country JSON.

Usage:
    python3 process_market_data.py                    # all countries in agent DB
    python3 process_market_data.py --country Thailand
    python3 process_market_data.py --list             # list available AEI files

Input files expected in data/raw/:
    aei_{country_slug}_{YYYY-MM}.xlsx   e.g. aei_thailand_2025-12.xlsx
    visa_grants_latest.xlsx             Home Affairs student visa data

Output:
    data/processed/market_size_{YYYY-MM}.json   (all countries, for display layer)
    data/processed/market_size_{YYYY-MM}.csv    (flat version for analysis)

--- HOW TO DOWNLOAD SOURCE FILES ---

AEI (commencements by nationality):
  1. Go to: https://www.education.gov.au/international-education-data-and-research/
             international-student-monthly-summary-and-data-tables
  2. Open the Tableau dashboard ("Pivot - Detailed" tab)
  3. Set the Nationality filter to e.g. "Thailand"
  4. Click Download → Data → Crosstab → Download
  5. Save as: data/raw/aei_thailand_2025-12.xlsx  (adjust month)
  6. Repeat for Nepal: data/raw/aei_nepal_2025-12.xlsx

Home Affairs (offshore visa grants):
  1. Go to: https://data.gov.au/data/dataset/student-visas
  2. Download most recent Excel file from "Student visa program" dataset
  3. Save as: data/raw/visa_grants_latest.xlsx
"""

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR  = Path(__file__).parent
RAW_DIR   = BASE_DIR / "data" / "raw"
PROC_DIR  = BASE_DIR / "data" / "processed"
DB_PATH   = BASE_DIR.parent / "data" / "agents.db"

# AEI nationality name → exact spelling used in AEI filter/file
AEI_NATIONALITY_MAP = {
    "Vietnam":      "Viet Nam",
    "China":        "China (People's Republic of)",
    "Hong Kong":    "Hong Kong (SAR of China)",
    "South Korea":  "Korea, Republic of",
    "Taiwan":       "Taiwan",
    # Most others match standard English name exactly
}

# Home Affairs citizenship spellings (where different from country name in DB)
HA_CITIZENSHIP_MAP = {
    "Vietnam":      "Viet Nam",
    "South Korea":  "Republic of Korea",
    "China":        "China (People's Republic of)",
    "Hong Kong":    "Hong Kong SAR",
    "Taiwan":       "Taiwan",
}

SECTORS_DISPLAY = ["Higher Education", "VET", "ELICOS", "Schools", "Non-award"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def country_slug(country: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", country.lower()).strip("_")


def yoy_pct(current, previous):
    if not previous or previous == 0 or current is None:
        return None
    return round(((current - previous) / previous) * 100, 1)


def find_aei_file(country: str) -> Path | None:
    """Find the most recent AEI file for this country."""
    slug = country_slug(country)
    files = sorted(RAW_DIR.glob(f"aei_{slug}_*.xlsx"), reverse=True)
    return files[0] if files else None


def aei_month_from_filename(path: Path) -> str:
    """Extract YYYY-MM from filename, return human-readable 'Month YYYY'."""
    m = re.search(r"(\d{4})-(\d{2})", path.name)
    if not m:
        return "Unknown"
    year, month = int(m.group(1)), int(m.group(2))
    return datetime(year, month, 1).strftime("%B %Y")


# ── AEI Parser ────────────────────────────────────────────────────────────────

def parse_aei_pivot(filepath: Path, country: str) -> dict:
    """
    Parse AEI_Pivot sheet from the downloaded Tableau export.

    Sheet layout (rows are 1-indexed):
      Rows 1–12:  Active filter values (Nationality, State, Sector etc)
      Row 15:     Measure label row  (Enrolments | Commencements)
      Row 16:     Year columns
      Rows 17–21: Higher Education | VET | Schools | ELICOS | Non-award
      Row 22:     Grand Total
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Try sheet named AEI_Pivot, fall back to first sheet
    if "AEI_Pivot" in wb.sheetnames:
        ws = wb["AEI_Pivot"]
    else:
        ws = wb.active
        print(f"    ⚠  Sheet 'AEI_Pivot' not found — using active sheet '{ws.title}'")

    # ── Read active filters (rows 1–12) ──
    filters = {}
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if row[0] and row[1]:
            filters[str(row[0]).strip()] = str(row[1]).strip()

    nationality_in_file = filters.get("Nationality", "(All)")
    print(f"    Nationality filter in file: {nationality_in_file}")

    # ── Find measure/year header rows ──
    # Scan for the row that contains "Enrolments" or "Commencements" as measure label
    # and the following row for year numbers. Tolerates layout shifts.
    measure_row_idx = None
    year_row_idx    = None

    for i, row in enumerate(ws.iter_rows(min_row=13, max_row=25, values_only=True), start=13):
        vals = [str(v).strip() if v else "" for v in row]
        if any("Enrolments" in v or "Commencements" in v for v in vals):
            measure_row_idx = i
            year_row_idx    = i + 1
            break

    if not measure_row_idx:
        # Fallback to spec positions
        measure_row_idx = 15
        year_row_idx    = 16
        print(f"    ⚠  Could not auto-detect header rows — using spec defaults (15/16)")

    # Read year headers
    year_row = list(ws.iter_rows(
        min_row=year_row_idx, max_row=year_row_idx, values_only=True
    ))[0]

    # Detect column ranges for enrolments vs commencements by reading measure row
    measure_row = list(ws.iter_rows(
        min_row=measure_row_idx, max_row=measure_row_idx, values_only=True
    ))[0]

    # Find column index where Commencements block starts
    enrol_cols   = []
    comm_cols    = []
    in_comm      = False
    for col_idx, val in enumerate(measure_row):
        v = str(val).strip() if val else ""
        if "Commencement" in v:
            in_comm = True
        if not in_comm and isinstance(year_row[col_idx], int):
            enrol_cols.append(col_idx)
        elif in_comm and isinstance(year_row[col_idx], int):
            comm_cols.append(col_idx)

    if not comm_cols:
        # Fallback: assume cols 1–7 = enrolments, 8–14 = commencements (spec layout)
        all_year_cols = [i for i, v in enumerate(year_row) if isinstance(v, int)]
        mid = len(all_year_cols) // 2
        enrol_cols = all_year_cols[:mid]
        comm_cols  = all_year_cols[mid:]

    enrol_years = [year_row[c] for c in enrol_cols if isinstance(year_row[c], int)]
    comm_years  = [year_row[c] for c in comm_cols  if isinstance(year_row[c], int)]

    # ── Read sector rows ──
    target_sectors = {s.lower(): s for s in SECTORS_DISPLAY + ["Grand Total"]}
    data = {}

    for row in ws.iter_rows(min_row=year_row_idx + 1, max_row=year_row_idx + 10, values_only=True):
        label = str(row[0]).strip() if row[0] else ""
        if label.lower() in target_sectors:
            canonical = target_sectors[label.lower()]
            data[canonical] = {
                "enrolments":    {yr: row[c] for yr, c in zip(enrol_years, enrol_cols)},
                "commencements": {yr: row[c] for yr, c in zip(comm_years,  comm_cols)},
            }

    wb.close()
    return {
        "nationality_filter": nationality_in_file,
        "filters": filters,
        "sectors": data,
        "comm_years": sorted(set(comm_years)),
        "enrol_years": sorted(set(enrol_years)),
    }


# ── Home Affairs Parser ───────────────────────────────────────────────────────

def parse_visa_offshore(filepath: Path, country: str) -> dict:
    """
    Parse Home Affairs student visa Excel for offshore grants by year.
    Returns {year: grant_count} for the given citizenship/country.
    """
    ha_country = HA_CITIZENSHIP_MAP.get(country, country)

    # Try to find the right sheet
    xl = pd.ExcelFile(filepath)
    sheet = None
    for name in xl.sheet_names:
        if any(kw in name.lower() for kw in ("grant", "visa", "data", "student")):
            sheet = name
            break
    if not sheet:
        sheet = xl.sheet_names[0]

    print(f"    Home Affairs sheet: '{sheet}'")
    df = pd.read_excel(filepath, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]

    # Find citizenship and offshore/onshore columns (case-insensitive)
    col_map = {c.lower(): c for c in df.columns}
    citizen_col  = next((col_map[k] for k in col_map if "citizen" in k), None)
    offshore_col = next((col_map[k] for k in col_map if "offshore" in k or "onshore" in k), None)
    year_col     = next((col_map[k] for k in col_map if k in ("year", "fin_year", "financial_year", "fy")), None)
    grants_col   = next((col_map[k] for k in col_map if "grant" in k), None)

    if not all([citizen_col, offshore_col, year_col, grants_col]):
        print(f"    ⚠  Could not identify required columns. Found: {list(df.columns)[:10]}")
        print(f"       Expected: citizenship, offshore/onshore, year, grants")
        return {}

    # Filter: country + offshore only
    mask = (
        df[citizen_col].astype(str).str.strip().str.lower() == ha_country.lower()
    ) & (
        df[offshore_col].astype(str).str.strip().str.lower() == "offshore"
    )
    filtered = df[mask]

    if filtered.empty:
        print(f"    ⚠  No offshore visa records found for '{ha_country}'")
        print(f"       Sample citizenships in file: {df[citizen_col].dropna().unique()[:8].tolist()}")
        return {}

    result = (
        filtered.groupby(year_col)[grants_col]
        .sum()
        .astype(int)
        .to_dict()
    )
    return {str(k): v for k, v in sorted(result.items())}


# ── Build country output ──────────────────────────────────────────────────────

def build_country_output(country: str, aei_data: dict, visa_data: dict, aei_month: str) -> dict:
    """Assemble the per-country JSON output block."""
    sectors = aei_data["sectors"]
    comm_years = sorted(aei_data["comm_years"])

    # Latest 3 years
    y3 = comm_years[-3:] if len(comm_years) >= 3 else comm_years

    def sector_block(sector_name: str) -> dict:
        s = sectors.get(sector_name, {}).get("commencements", {})
        vals = {str(y): (s.get(y) or 0) for y in y3}
        # YoY on last two years
        yrs = sorted(vals.keys())
        yoy = yoy_pct(vals.get(yrs[-1], 0), vals.get(yrs[-2], 0)) if len(yrs) >= 2 else None
        return {**vals, "yoy_pct": yoy}

    # Total commencements = Grand Total row, or sum of sectors
    if "Grand Total" in sectors:
        total_block = sector_block("Grand Total")
    else:
        totals = {}
        for yr in y3:
            totals[str(yr)] = sum(
                (sectors.get(s, {}).get("commencements", {}).get(yr) or 0)
                for s in SECTORS_DISPLAY
            )
        yrs = sorted(totals.keys())
        yoy = yoy_pct(totals.get(yrs[-1], 0), totals.get(yrs[-2], 0)) if len(yrs) >= 2 else None
        total_block = {**totals, "yoy_pct": yoy}

    # Other = all sectors minus HE + VET + ELICOS
    other_sectors = [s for s in sectors if s not in ("Higher Education", "VET", "ELICOS", "Grand Total")]
    other_vals = {}
    for yr in y3:
        other_vals[str(yr)] = sum(
            (sectors.get(s, {}).get("commencements", {}).get(yr) or 0)
            for s in other_sectors
        )
    other_yrs = sorted(other_vals.keys())
    other_yoy = yoy_pct(
        other_vals.get(other_yrs[-1], 0),
        other_vals.get(other_yrs[-2], 0)
    ) if len(other_yrs) >= 2 else None

    # Offshore visa grants — last 3 years
    visa_years = sorted(visa_data.keys())[-3:] if visa_data else []
    visa_block = {y: visa_data[y] for y in visa_years}
    if len(visa_years) >= 2:
        visa_block["yoy_pct"] = yoy_pct(
            visa_data.get(visa_years[-1], 0),
            visa_data.get(visa_years[-2], 0)
        )

    return {
        "country": country,
        "aei_month": aei_month,
        "commencements": {
            "total": total_block,
            "by_sector": {
                "Higher Education": sector_block("Higher Education"),
                "VET":              sector_block("VET"),
                "ELICOS":           sector_block("ELICOS"),
                "Other":            {**{str(y): other_vals[str(y)] for y in y3}, "yoy_pct": other_yoy},
            }
        },
        "offshore_visa_grants": visa_block,
        "sources": {
            "aei_url":         "https://www.education.gov.au/international-education-data-and-research/international-student-monthly-summary-and-data-tables",
            "homeaffairs_url": "https://data.gov.au/data/dataset/student-visas",
        }
    }


# ── Display block renderer ────────────────────────────────────────────────────

def render_display_block(out: dict) -> str:
    """Render the text market size block for terminal/report output."""
    c  = out["commencements"]
    t  = c["total"]
    bs = c["by_sector"]

    # Get year keys (exclude yoy_pct)
    yrs = sorted(k for k in t if k != "yoy_pct")

    def arrow(yoy):
        if yoy is None: return "—"
        if yoy > 5:  return f"↑ +{yoy}%"
        if yoy < -5: return f"↓ {yoy}%"
        return f"→ {yoy:+.1f}%"

    def fmt(v):
        if v is None: return "    —"
        return f"{v:>6,}"

    lines = []
    lines.append(f"\nMARKET SIZE — {out['country'].upper():<20} [AEI {out['aei_month']}]")
    lines.append("")
    lines.append(f"  {'Commencements':<22} " + "  ".join(f"{y:>7}" for y in yrs) + "  Trend")
    lines.append("  " + "─" * 65)

    for sector_label in ["Higher Education", "VET", "ELICOS", "Other"]:
        s = bs.get(sector_label, {})
        row_vals = "  ".join(fmt(s.get(y)) for y in yrs)
        lines.append(f"  {sector_label:<22} {row_vals}  {arrow(s.get('yoy_pct'))}")

    lines.append("  " + "─" * 65)
    total_vals = "  ".join(fmt(t.get(y)) for y in yrs)
    lines.append(f"  {'Total':<22} {total_vals}  {arrow(t.get('yoy_pct'))}")

    # Visa grants
    vg = out.get("offshore_visa_grants", {})
    if vg:
        visa_yrs = sorted(k for k in vg if k != "yoy_pct")
        visa_str = "  →  ".join(f"{y}: {vg[y]:,}" for y in visa_yrs)
        lines.append(f"\n  Offshore visa grants    {visa_str}")

    lines.append(
        "\n  Latest monthly figures →\n"
        "  Australian Dept of Education — International Student Monthly Summary\n"
        "  education.gov.au/international-education-data-and-research/"
        "international-student-monthly-summary-and-data-tables"
    )
    lines.append(
        "\n  Data: Australian Department of Education | Department of Home Affairs"
    )
    return "\n".join(lines)


# ── CSV export ────────────────────────────────────────────────────────────────

def build_csv_rows(all_countries: list[dict]) -> list[dict]:
    rows = []
    for out in all_countries:
        country = out["country"]
        aei_month = out["aei_month"]
        t = out["commencements"]["total"]
        bs = out["commencements"]["by_sector"]
        vg = out.get("offshore_visa_grants", {})
        yrs = sorted(k for k in t if k != "yoy_pct")

        for sector in list(bs.keys()) + ["Total"]:
            s = t if sector == "Total" else bs.get(sector, {})
            row = {"country": country, "aei_month": aei_month, "sector": sector}
            for yr in yrs:
                row[f"commencements_{yr}"] = s.get(yr)
            row["yoy_pct"] = s.get("yoy_pct")
            rows.append(row)

        if vg:
            v_yrs = sorted(k for k in vg if k != "yoy_pct")
            row = {"country": country, "aei_month": aei_month, "sector": "Offshore Visa Grants"}
            for yr in v_yrs:
                row[f"commencements_{yr}"] = vg.get(yr)
            row["yoy_pct"] = vg.get("yoy_pct")
            rows.append(row)
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────

def get_countries_from_db() -> list[str]:
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT DISTINCT country FROM agent_social ORDER BY country"
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--country", help="Single country to process")
    parser.add_argument("--list", action="store_true", help="List available AEI files")
    args = parser.parse_args()

    if args.list:
        files = sorted(RAW_DIR.glob("aei_*.xlsx"))
        if not files:
            print("No AEI files found in data/raw/")
        for f in files:
            print(f"  {f.name}")
        return

    countries = [args.country] if args.country else get_countries_from_db()

    # Load Home Affairs file once (covers all countries)
    visa_file = next(iter(sorted(RAW_DIR.glob("visa_grants*.xlsx"), reverse=True)), None)
    if not visa_file:
        print("⚠  No Home Affairs visa file found (data/raw/visa_grants_latest.xlsx)")
        print("   Download from: https://data.gov.au/data/dataset/student-visas")
        print("   Continuing without visa data.\n")

    all_outputs = []
    aei_month_latest = None

    for country in countries:
        print(f"\n── {country} ─────────────────")

        aei_file = find_aei_file(country)
        if not aei_file:
            print(f"  ⚠  No AEI file found for {country}")
            print(f"     Expected: data/raw/aei_{country_slug(country)}_YYYY-MM.xlsx")
            print(f"     Download instructions: see top of this script")
            continue

        aei_month = aei_month_from_filename(aei_file)
        aei_month_latest = aei_month
        print(f"  📊 AEI: {aei_file.name} ({aei_month})")

        try:
            aei_data = parse_aei_pivot(aei_file, country)
        except Exception as e:
            print(f"  ✗ AEI parse error: {e}")
            continue

        visa_data = {}
        if visa_file:
            print(f"  🛂 Visa: {visa_file.name}")
            try:
                visa_data = parse_visa_offshore(visa_file, country)
                if visa_data:
                    print(f"    Found {len(visa_data)} years: {list(sorted(visa_data.keys()))}")
            except Exception as e:
                print(f"  ✗ Visa parse error: {e}")

        out = build_country_output(country, aei_data, visa_data, aei_month)
        all_outputs.append(out)
        print(render_display_block(out))

    if not all_outputs:
        print("\nNo data processed. Download AEI files and try again.")
        return

    # Save JSON
    month_tag = re.sub(r"[^0-9-]", "", all_outputs[0]["aei_month"]) or "latest"
    # Convert "December 2025" → "2025-12"
    try:
        dt = datetime.strptime(all_outputs[0]["aei_month"], "%B %Y")
        month_tag = dt.strftime("%Y-%m")
    except Exception:
        month_tag = "latest"

    json_path = PROC_DIR / f"market_size_{month_tag}.json"
    with open(json_path, "w") as f:
        json.dump(all_outputs, f, indent=2, ensure_ascii=False)
    print(f"\n✅ JSON → {json_path}")

    # Save CSV
    csv_path = PROC_DIR / f"market_size_{month_tag}.csv"
    rows = build_csv_rows(all_outputs)
    if rows:
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        print(f"✅ CSV  → {csv_path}")


if __name__ == "__main__":
    main()
