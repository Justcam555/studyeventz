#!/usr/bin/env python3
"""
Agent Database Query Tool

Usage:
    python query.py agents                              # All agents
    python query.py agents --country "China"           # By country
    python query.py agents --university "Monash"       # By university
    python query.py agents --email "@gmail.com"        # Email pattern
    python query.py agents --export agents.xlsx        # Export to Excel

    python query.py stats                              # Summary statistics
    python query.py stats --by country                 # Stats grouped by country
    python query.py stats --by university              # Stats by university

    python query.py coverage                           # Which unis have data
    python query.py search "education group"           # Full-text search
"""

import argparse
import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).parent / "data" / "agents.db"


def get_conn():
    if not DB_PATH.exists():
        print("❌  Database not found. Run scrape.py first.")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ─── Queries ──────────────────────────────────────────────────────────────────

def query_agents(conn, country=None, university=None, email_pattern=None,
                 has_email=False, has_website=False, limit=None):
    sql = """
        SELECT
            u.name          AS university,
            a.company_name,
            a.contact_name,
            a.country,
            a.region,
            a.city,
            a.email,
            a.phone,
            a.website,
            a.address,
            a.scraped_at
        FROM agents a
        JOIN universities u ON u.id = a.university_id
        WHERE 1=1
    """
    params = []

    if country:
        sql += " AND LOWER(a.country) LIKE LOWER(?)"
        params.append(f"%{country}%")
    if university:
        sql += " AND LOWER(u.name) LIKE LOWER(?)"
        params.append(f"%{university}%")
    if email_pattern:
        sql += " AND LOWER(a.email) LIKE LOWER(?)"
        params.append(f"%{email_pattern}%")
    if has_email:
        sql += " AND a.email IS NOT NULL AND a.email != ''"
    if has_website:
        sql += " AND a.website IS NOT NULL AND a.website != ''"

    sql += " ORDER BY a.country, u.name, a.company_name"
    if limit:
        sql += f" LIMIT {int(limit)}"

    return pd.read_sql_query(sql, conn, params=params)


def query_stats(conn, group_by="country"):
    if group_by == "country":
        sql = """
            SELECT
                COALESCE(a.country, '(unknown)') AS country,
                COUNT(DISTINCT u.id)             AS universities,
                COUNT(a.id)                      AS total_agents,
                COUNT(a.email)                   AS with_email,
                COUNT(a.website)                 AS with_website,
                COUNT(a.phone)                   AS with_phone
            FROM agents a
            JOIN universities u ON u.id = a.university_id
            GROUP BY 1
            ORDER BY total_agents DESC
        """
    elif group_by == "university":
        sql = """
            SELECT
                u.name                           AS university,
                COUNT(a.id)                      AS total_agents,
                COUNT(DISTINCT a.country)        AS countries_covered,
                COUNT(a.email)                   AS with_email,
                u.scrape_status,
                u.last_scraped
            FROM universities u
            LEFT JOIN agents a ON a.university_id = u.id
            GROUP BY u.id
            ORDER BY total_agents DESC
        """
    elif group_by == "country_university":
        sql = """
            SELECT
                COALESCE(a.country, '(unknown)') AS country,
                u.name                           AS university,
                COUNT(a.id)                      AS agents
            FROM agents a
            JOIN universities u ON u.id = a.university_id
            GROUP BY 1, 2
            ORDER BY 1, 3 DESC
        """
    else:
        raise ValueError(f"Unknown group_by: {group_by}")

    return pd.read_sql_query(sql, conn)


def query_coverage(conn):
    return pd.read_sql_query("""
        SELECT
            u.name,
            u.agent_page_url,
            u.scrape_status,
            u.last_scraped,
            COUNT(a.id) AS agents_in_db,
            u.status_note
        FROM universities u
        LEFT JOIN agents a ON a.university_id = u.id
        GROUP BY u.id
        ORDER BY agents_in_db DESC, u.name
    """, conn)


def search_agents(conn, term):
    return pd.read_sql_query("""
        SELECT
            u.name AS university,
            a.company_name, a.contact_name,
            a.country, a.city,
            a.email, a.phone, a.website,
            a.raw_text
        FROM agents a
        JOIN universities u ON u.id = a.university_id
        WHERE LOWER(a.raw_text)     LIKE LOWER(?)
           OR LOWER(a.company_name) LIKE LOWER(?)
           OR LOWER(a.contact_name) LIKE LOWER(?)
           OR LOWER(a.email)        LIKE LOWER(?)
        ORDER BY u.name, a.company_name
    """, conn, params=[f"%{term}%"] * 4)


# ─── Export ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
BORDER      = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def export_excel(df, path, sheet_name="Agents"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Headers
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            # Clickable URLs
            if val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

    # Auto-width
    for ci, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            *(len(str(v)) for v in df[col] if v is not None)
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"✅  Exported {len(df)} rows → {path}")


# ─── Print helpers ────────────────────────────────────────────────────────────

def print_df(df, max_rows=50):
    if df.empty:
        print("  (no results)")
        return
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 180)
    pd.set_option("display.max_colwidth", 40)
    print(df.head(max_rows).to_string(index=False))
    if len(df) > max_rows:
        print(f"  … {len(df) - max_rows} more rows (use --export to get all)")
    print(f"\n  Total: {len(df)} rows")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Query the AU Agent database")
    sub = parser.add_subparsers(dest="cmd")

    # agents
    p_agents = sub.add_parser("agents", help="Query agent records")
    p_agents.add_argument("--country",    help="Filter by country")
    p_agents.add_argument("--university", help="Filter by university name")
    p_agents.add_argument("--email",      help="Filter by email pattern")
    p_agents.add_argument("--has-email",  action="store_true")
    p_agents.add_argument("--has-website",action="store_true")
    p_agents.add_argument("--limit",      type=int)
    p_agents.add_argument("--export",     metavar="FILE.xlsx")

    # stats
    p_stats = sub.add_parser("stats", help="Summary statistics")
    p_stats.add_argument("--by", choices=["country", "university", "country_university"],
                         default="country")
    p_stats.add_argument("--export", metavar="FILE.xlsx")

    # coverage
    p_cov = sub.add_parser("coverage", help="Scrape coverage by university")
    p_cov.add_argument("--export", metavar="FILE.xlsx")

    # search
    p_search = sub.add_parser("search", help="Full-text search across all fields")
    p_search.add_argument("term", help="Search term")
    p_search.add_argument("--export", metavar="FILE.xlsx")

    args = parser.parse_args()

    if not args.cmd:
        parser.print_help()
        return

    conn = get_conn()

    if args.cmd == "agents":
        df = query_agents(conn,
                          country=args.country,
                          university=args.university,
                          email_pattern=args.email,
                          has_email=args.has_email,
                          has_website=args.has_website,
                          limit=args.limit)
        if args.export:
            export_excel(df, args.export, "Agents")
        else:
            print_df(df)

    elif args.cmd == "stats":
        df = query_stats(conn, group_by=args.by)
        if args.export:
            export_excel(df, args.export, f"Stats by {args.by}")
        else:
            print_df(df, max_rows=100)

    elif args.cmd == "coverage":
        df = query_coverage(conn)
        if args.export:
            export_excel(df, args.export, "Coverage")
        else:
            print_df(df, max_rows=50)

    elif args.cmd == "search":
        df = search_agents(conn, args.term)
        if args.export:
            export_excel(df, args.export, "Search Results")
        else:
            print_df(df)

    conn.close()


if __name__ == "__main__":
    main()
